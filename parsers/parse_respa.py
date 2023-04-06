from openpyxl import load_workbook
from objects.result import Result
import tools

def parse_results(year):
    """
    Parses the results of the Republican Subject Olympiad (RESPA) from an Excel file for the given year.

    Args:
        year (int): The year of the Olympiad.

    Returns:
        dict: A dictionary containing the parsed results for each grade. The keys of the dictionary are the grade
        levels (9, 10, and 11), and the values are lists of Result objects representing the results for each student
        in the corresponding grade level.

    """
    bygrade = {}
    wb = load_workbook(f"data/results/respa/{year}.xlsx", data_only=True)
    for grade in (9, 10, 11,):
        ws = wb[f"{grade} класс"]
        row = 3
        while True:
            re = Result()
            re.name = ws['A'+str(row)].value
            re.total = 0
            re.only_total = False
            re.olymp_name = "respa"
            if re.name is None:
                break
            col = 'C'
            while True:
                if '№' in ws[col+'2'].value:
                    num = int(ws[col+'2'].value.split('№')[1])
                    maxscore = int(ws[col+'1'].value)
                    res = float(ws[col+str(row)].value)
                    setattr(re, f"problem{num}", {'abs': res, 'max': maxscore, 'rel': res/maxscore})
                    re.total += res
                    # print(getattr(re, f"problem{num}"))
                    re.numprobs = num
                else:
                    break
                col = tools.getNextCol(col)
            bygrade.setdefault(grade, []).append(re)
            row += 1
    return bygrade


if __name__ == "__main__":
    parse_results(2023)