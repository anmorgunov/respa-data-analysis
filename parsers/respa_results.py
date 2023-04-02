from openpyxl import load_workbook
from objects.result import Result
import tools

def parse_results(year):
    bygrade = {}
    wb = load_workbook(f"data/results/respa/{year}.xlsx", data_only=True)
    for grade in (9, 10, 11,):
        ws = wb[f"{grade} класс"]
        row = 3
        while True:
            re = Result()
            re.name = ws['A'+str(row)].value
            if re.name is not None:
                col = 'C'
                while True:
                    if '№' in ws[col+'2'].value:
                        num = int(ws[col+'2'].value.split('№')[1])
                        maxscore = int(ws[col+'1'].value)
                        res = float(ws[col+str(row)].value)
                        setattr(re, f"problem{num}", {'abs': res, 'max': maxscore, 'rel': res/maxscore})
                        # print(getattr(re, f"problem{num}"))
                        re.numprobs = num
                    else:
                        break
                    col = tools.getNextCol(col)
            else:
                break
            bygrade.setdefault(grade, []).append(re)
            row += 1
    return bygrade


if __name__ == "__main__":
    parse_results(2023)